import json
from translate import *
import shutil
import os
import jsonpatch
from openpyxl import load_workbook
from openpyxl import workbook
def get_json_from_class(fromlan,classs):
    if fromlan=='jp' or fromlan=='kor':
        return classs.json_en
    else:return classs.json_kr
def get_json_target_class(fromlan,classs):
    if fromlan=='en':
        return classs.json_en
    elif fromlan=='kor':
        return classs.json_kr
    else:return classs.json_jp
class transing_cache:
    name_path=''
    path_en=''
    path_llc=''
    path_kr=''
    path_jp=''
    json_en=''
    json_jp=''
    json_kr=''
    json_llc=''
def log(path,text,trans_ini):
    if trans_ini.log:
        with open(path,'a') as f:
            f.write(text+'\n')
def trans_try(json_from,json_to,trans_ini,path_log,trans_cache):
    diff=jsonpatch.JsonPatch.from_diff(json_from,json_to)
    for i in diff:
        if i['op']=='replace':
            after=translate_final(fromlang,'zh',i['value'],inter,appid,appkey)
            print(after)
            i['value']=after
    returnz=jsonpatch.apply_patch(json_for,diff)
    return returnz
def read_sheet_content(file_path, sheet_name):
        # 加载工作簿
        wb = load_workbook(file_path)
        # 获取指定的 sheet
        ws = wb[sheet_name]
        # 读取所有行的内容
        sheet_data = []
        for row in ws.iter_rows(values_only=True):  # values_only=True 表示只获取单元格的值
            sheet_data.append(list(row))
        return sheet_data  
def create_sheet_in_excel(file_path, new_sheet_name):
    try:
        # 尝试加载现有的工作簿
        wb = load_workbook(file_path)
    except FileNotFoundError:
        # 如果文件不存在，则创建一个新的工作簿
        wb = Workbook()
        # 删除默认创建的 'Sheet' 表
        default_sheet = wb.active
        wb.remove(default_sheet)
    # 检查是否已经存在同名 sheet
    if new_sheet_name in wb.sheetnames:
        return
    wb.create_sheet(title=new_sheet_name)
    wb.save(file_path)
def append_row_to_sheet(file_path, sheet_name, row_data):
    # 加载工作簿和指定的 sheet
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    # 在 sheet 的末尾追加一行数据
    ws.append(row_data)
    # 保存更改
    wb.save(file_path)
def get_json_from(fromlan,en,kor):
    if fromlan=='jp' or fromlan=='kor':
        return en
    else:return kor
def get_json_target(fromlan,en,kor,jp):
    if fromlan=='en':
        return en
    elif fromlan=='kor':
        return kor
    else:return jp
def mymovefile(srcfile,dstpath,with_let,in_folder):                       # 移动函数
    fpath,fname=os.path.split(srcfile)             # 分离文件名和路径
    if (not in_folder=='') and (not os.path.exists(dstpath+'\\'+in_folder[:-1])):
        os.makedirs(dstpath+'\\'+in_folder[:-1])
    if with_let:
        shutil.copy(srcfile, dstpath +'\\'+ in_folder + fname[3:])
    else:
        shutil.copy(srcfile, dstpath +'\\'+ in_folder +fname)          # 移动文件
def make_id_list(list):
    result=[]
    for i in list:
        if not i=={}:
            try:
                result.append(i['id'])
            except:
                result.append(None)
    return result
def make_id_list_true(list):
    result=[]
    for i in list:
        try:
            result.append(i['id'])
        except:
            result.append(None)
    return result
def doing_the_trans(value,trans_ini):
    if not trans_ini.using_self:
        ret=translate_final(value,trans_ini)
    else:
        ret=trans_ini.trans_modue.trans_selfdir(value,trans_ini.key)
    if ret=='error':
        return value
    return ret
def all_trans(json_for,json_to,trans_ini):

    diff=jsonpatch.JsonPatch.from_diff(json_for,json_to)
    for i in diff:
        if i['op']=='replace':
            after=doing_the_trans(fromlang,i['value'],formal_input,using_self,)
            print(after)
            i['value']=after
    returnz=jsonpatch.apply_patch(json_for,diff)
    return returnz
def one_trans(path_limbus,path_llc,name,path_output,trans_ini):
    None
def find_type(path_limbus,path_llc,name,path_output,trans_ini):
    name_path=(os.path.split(name))[0]
    if not name_path=='':
        name_path=name_path+'\\'
    name_file=(os.path.split(name))[1]
    path_en=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\en\\'+name_path+'EN_'+name_file
    path_kor=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr\\'+name_path+'KR_'+name_file
    path_jp=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\jp\\'+name_path+'JP_'+name_file
    path_llc=path_llc+'\\'+name
    trans_cache=transing_cache()
    #name_path=name_path,path_en=path_en,path_kor=path_kor,path_jp=path_jp,path_llc=path_llc
    trans_cache.name_path=name_path
    trans_cache.path_en=path_en
    trans_cache.path_kr=path_kor
    trans_cache.path_jp=path_jp
    trans_cache.path_llc=path_llc
    if name_file=='E000X.json' :
        return 'not_need',trans_cache
    with open(path_en, 'rb') as file:
        json_en=json.load(file) 
        trans_cache.json_en=json_en
    llc_exist=True
    if json_en=={}:
        return 'empety',trans_cache
    if os.path.exists(path_llc):
        with open(path_llc, 'rb') as file:
            json_llc=json.load(file)
            trans_cache.json_llc=json_llc
    else:
        llc_exist=False
    try:
        list_en=json_en['dataList']
    except:
        return 'json_error',trans_cache
    if list_en==[] or list_en==[{}]:
        return 'empety',trans_cache
    id_en=make_id_list(list_en)
    if llc_exist:
        list_llc=json_llc['dataList']
        id_llc=make_id_list(list_llc)
        if id_en==id_llc:
            return 'correct_llc',trans_cache
        else:
            just_none=True
            for i in id_en:
                if not i==None:
                    just_none=False
            if just_none and len(id_en)==len(id_llc):
                return 'no_id_correct_llc',trans_cache
            else:
                if just_none:
                    return 'no_id_but_diff',trans_cache
                elif len(id_en)<len(id_llc):
                    for i in id_en:
                        if not i in id_llc:return 'normal_diff_longer_llc',trans_cache
                    return 'correct_longer_llc',trans_cache
                else:
                    return 'normal_diff',trans_cache
    else:
        return 'not_llc_exist',trans_cache
    return 'error'
def cp_trans_new(path_limbus,path_llc,name,path_output,trans_ini,path_log):
    type_file,trans_cache=find_type(path_limbus,path_llc,name,path_output,trans_ini)
    log(path_log,'\nfile_name:'+name+'type_file:'+type_file,trans_ini)
    if type_file=='not_need' or type_file=='empety':
        if trans_ini.exc:
            append_row_to_sheet(path_output,'have_trans',name)
        else:
            mymovefile(trans_cache.path_en,path_output,True,trans_cache.name_path)
    elif type_file=='correct_llc' or type_file=='correct_longer_llc' or type_file=='no_id_correct_llc':
        if trans_ini.exc:
            append_row_to_sheet(path_output,'have_trans',name)
        else:
            mymovefile(trans_cache.path_llc,path_output,True,trans_cache.name_path)
    elif type_file=='json_error':
        print('错误:json_error,无法找到datalist\n文件名称:'+name)
        while True:
            if trans_ini.json_error=='for':
                mymovefile(trans_cache.path_en,path_output,True,trans_cache.name_path)
                log(path_log,'已使用原文件处理json文件错误',trans_ini)
                break
            elif trans_ini.json_error=='skip':
                log(path_log,'已跳过json文件错误',trans_ini)
                break
            elif trans_ini.json_error=='try':
                trans_try()
                break
            else:
                ask=input('请输入解决方案(for/llc/skip/try)?以帮助')
                if ask=='afor':
                    trans_ini.json_error='for'
                    log(path_log,'使用原文件处理json文件错误',trans_ini)
                elif ask=='for':
                    mymovefile(trans_cache.path_en,path_output,False,trans_cache.name_path)
                    log(path_log,'已使用原文件处理json文件错误',trans_ini)
                    break
                elif ask=='askip':
                    trans_ini.json_error='skip'
                    log(path_log,'已设置跳过json文件错误',trans_ini)
                elif ask=='skip':
                    log(path_log,'已跳过json文件错误',trans_ini)
                    break
                elif ask=='llc':
                    if os.path.exists(trans_cache.path_llc):
                        mymovefile(trans_cache.path_llc,path_output,False,trans_cache.name_path)
                        log(path_log,'已移动零协文件',trans_ini)
                        break
                    else:
                        print('零协文件不存在')
                elif ask=='try':
                    trans_try()
                    log(path_log,'尝试使用不稳定方案进行通篇汉化',trans_ini)
                    break
                elif ask=='atry':
                    trans_ini.json_error='try'
                    log(path_log,'设置使用不稳定方案进行通篇汉化',trans_ini)
                elif ask=='?' or ask=='？':
                    print('在解析文件内容时出现了错误')
                    print('请选择一种解决方案')
                    print('for:将原文件输出')
                    print('skip:跳过该文件(可能导致汉化错误)')
                    print('try:尝试使用不稳定方案进行通篇汉化')
                    print('llc:使用LLC汉化(可能不存在)')
                    print('在输入选项前加入a(无法作用于llc)以应用于本次翻译所有此项操作')
    elif type_file=='no_id_but_diff':
        while True:
            ask=input('文件'+name+'无id且与llc长度不一,无法进行常规翻译，选择方案(for/skip/try/llc/?)')
            if ask=='for':
                mymovefile(trans_cache.path_en,path_output,True,trans_cache.name_path)
                log(path_log,'已使用原文件方案',trans_ini)
                break
            elif ask=='skip':
                log(path_log,'已跳过该文件',trans_ini)
                break
            elif ask=='try':
                trans_try()
                log(path_log,'尝试使用不稳定方案进行通篇汉化',trans_ini)
                break
            elif ask=='llc':
                if os.path.exists(trans_cache.path_llc):
                    mymovefile(trans_cache.path_llc,path_output,False,trans_cache.name_path)
                    log(path_log,'已移动零协文件',trans_ini)
                else:
                    print('零协文件不存在')
            elif ask=='?' or ask=='？':
                print('请选择一种解决方案')
                print('for:将原文件输出')
                print('skip:跳过该文件(可能导致汉化错误)')
                print('try:尝试使用不稳定方案进行通篇汉化')
                print('llc:使用LLC汉化(可能不存在)')
    elif type_file=='normal_diff':
        if trans_ini.exc:
            sheet_name=name.replace('.json','').replace('\\','_')
            create_sheet_in_excel(path_output,sheet_name)
        final_list=[]
        with open(trans_cache.path_kr, 'rb') as file:
            trans_cache.json_kr=json.load(file)
        with open(trans_cache.path_jp, 'rb') as file:
            trans_cache.json_jp=json.load(file)
        list_llc=trans_cache.json_llc['dataList']
        id_llc_true=make_id_list_true(trans_cache.json_llc)
        id_target_true=make_id_list_true(get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'])
        id_for_true=make_id_list_true(get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'])
        if not (trans_ini.exc or trans_ini.team_trans):
            cache_trans_for_list=[]
            cache_trans_cn_list=[]
            diff_list=[]
            for i in id_target_true:
                id_in_llc=i in id_llc_true
                if id_in_llc:
                    final_list.append(list_llc[id_llc_true.index(i)])
                else:
                    json_for=get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'][id_for_true.index(i)]
                    json_target=get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'][id_target_true.index(i)]
                    diff=jsonpatch.JsonPatch.from_diff(json_for,json_target)
                    if trans_ini.cache_trans:
                        for im in diff:
                            if im['op']=='replace':
                                if im['value'] in cache_trans_for_list:
                                    after=cache_trans_cn_list[cache_trans_for_list.index(im['value'])]
                                else:
                                    cache_trans_for_list.append(im['value'])
                                    after=doing_the_trans(im['value'],trans_ini)
                                    cache_trans_cn_list.append(after)
                                print(after)
                                im['value']=after
                        final_list.append(jsonpatch.apply_patch(json_for,diff))
                    else:
                        for im in diff:
                            if im['op']=='replace':
                                after=doing_the_trans(im['value'],trans_ini)
                                print(after)
                                im['value']=after
                        final_list.append(jsonpatch.apply_patch(json_for,diff))
            ret={}
            ret['dataList']=final_list
            with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                json.dump(ret,f,ensure_ascii=False,indent=4)
        elif trans_ini.team_trans:
            need_trans_id=[]
            need_trans=[]
            json_for_list=[]
            diff_list=[]
            for i in id_target_true:
                id_in_llc=i in id_llc_true
                if id_in_llc:
                    need_trans_id.append(None)
                    need_trans.append(None)
                    json_for_list.append(None)
                    diff_list.append(None)
                    #final_list.append(list_llc[id_llc_true.index(i)])
                else:
                    need_trans_id.append(i)
                    json_for=get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'][id_for_true.index(i)]
                    json_target=get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'][id_target_true.index(i)]
                    diff=jsonpatch.JsonPatch.from_diff(json_for,json_target)
                    diff_list.append(diff)
                    json_for_list.append(json_for)
                    need_trans_in_id=[]
                    for im in diff:
                        if im['op']=='replace':
                            need_trans_in_id.append(im['value'])
                            #after=doing_the_trans(im['value'],trans_ini)
                            #print(after)
                            #im['value']=after
                    need_trans.append(need_trans_in_id)
                    #final_list.append(jsonpatch.apply_patch(json_for,diff))
            team_trans_list=[]
            for num in range(len(need_trans)):
                if need_trans[num] is not None:
                    for val in need_trans[num]:
                        team_trans_list.append(val)
            if trans_ini.using_self:
                team_trans_list=trans_ini.trans_modue.self_trans(team_trans_list)
            else:
                team_cn_list=trans_team(team_trans_list,trans_ini)
            if not len(team_cn_list)==len(team_trans_list):
                print('翻译错误')
                input('按回车继续')
                raise RuntimeError('teamtrans列表长度错误')
            print(team_cn_list)
            ok_trans=need_trans.copy()
            cn_time=0
            for num in range(len(ok_trans)):
                if ok_trans[num] is not None:
                    for val in ok_trans[num]:
                        ok_trans[num]=team_cn_list[cn_time]
                        cn_time+=1
            for i in range(len(need_trans_id)):
                if ok_trans[i] is not None:
                    ok_time=0
                    for im in diff_list[i]:
                        if im['op']=='replace':
                            im['value']=ok_trans[i][ok_time]
                            ok_time+=1
                    final_list.append(jsonpatch.apply_patch(json_for,diff_list[i]))
                else:
                    final_list.append(final_list.append(list_llc[id_llc_true.index(need_trans_id[i])]))
            ret={}
            ret['dataList']=final_list
            with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                json.dump(ret,f,ensure_ascii=False,indent=4)
    elif type_file=='not_llc_exist':
        sheet_name=name.replace('.json','').replace('\\','_')
        if trans_ini.exc:
            create_sheet_in_excel(path_output,sheet_name)
        final_list=[]
        with open(trans_cache.path_kr, 'rb') as file:
            trans_cache.json_kr=json.load(file)
        with open(trans_cache.path_jp, 'rb') as file:
            trans_cache.json_jp=json.load(file)
        list_llc=[]
        id_llc_true=[]
        id_target_true=make_id_list_true(get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'])
        id_for_true=make_id_list_true(get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'])
        if not (trans_ini.exc or trans_ini.team_trans):
            cache_trans_for_list=[]
            cache_trans_cn_list=[]
            diff_list=[]
            for i in id_target_true:
                id_in_llc=i in id_llc_true
                if id_in_llc:
                    final_list.append(list_llc[id_llc_true.index(i)])
                else:
                    json_for=get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'][id_for_true.index(i)]
                    json_target=get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'][id_target_true.index(i)]
                    diff=jsonpatch.JsonPatch.from_diff(json_for,json_target)
                    if trans_ini.cache_trans:
                        for im in diff:
                            if im['op']=='replace':
                                if im['value'] in cache_trans_for_list:
                                    after=cache_trans_cn_list[cache_trans_for_list.index(im['value'])]
                                else:
                                    cache_trans_for_list.append(im['value'])
                                    after=doing_the_trans(im['value'],trans_ini)
                                    cache_trans_cn_list.append(after)
                                print(after)
                                im['value']=after
                        final_list.append(jsonpatch.apply_patch(json_for,diff))
                    else:
                        for im in diff:
                            if im['op']=='replace':
                                after=doing_the_trans(im['value'],trans_ini)
                                print(after)
                                im['value']=after
                        final_list.append(jsonpatch.apply_patch(json_for,diff))
            ret={}
            ret['dataList']=final_list
            with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                json.dump(ret,f,ensure_ascii=False,indent=4)
        elif trans_ini.team_trans:
            need_trans_id=[]
            need_trans=[]
            json_for_list=[]
            diff_list=[]
            for i in id_target_true:
                id_in_llc=i in id_llc_true
                if id_in_llc:
                    need_trans_id.append(None)
                    need_trans.append(None)
                    json_for_list.append(None)
                    diff_list.append(None)
                    #final_list.append(list_llc[id_llc_true.index(i)])
                else:
                    need_trans_id.append(i)
                    json_for=get_json_from_class(trans_ini.lang_in,trans_cache)['dataList'][id_for_true.index(i)]
                    json_target=get_json_target_class(trans_ini.lang_in,trans_cache)['dataList'][id_target_true.index(i)]
                    diff=jsonpatch.JsonPatch.from_diff(json_for,json_target)
                    diff_list.append(diff)
                    json_for_list.append(json_for)
                    need_trans_in_id=[]
                    for im in diff:
                        if im['op']=='replace':
                            need_trans_in_id.append(im['value'])
                            #after=doing_the_trans(im['value'],trans_ini)
                            #print(after)
                            #im['value']=after
                    need_trans.append(need_trans_in_id)
                    #final_list.append(jsonpatch.apply_patch(json_for,diff))
            team_trans_list=[]
            for num in range(len(need_trans)):
                if need_trans[num] is not None:
                    for val in need_trans[num]:
                        team_trans_list.append(val)
            if trans_ini.using_self:
                team_trans_list=trans_ini.trans_modue.self_trans(team_trans_list)
            else:
                team_cn_list=trans_team(team_trans_list,trans_ini)
            if not len(team_cn_list)==len(team_trans_list):
                print('翻译错误')
                input('按回车继续')
                raise RuntimeError('teamtrans列表长度错误')
            print(team_cn_list)
            ok_trans=need_trans.copy()
            cn_time=0
            for num in range(len(ok_trans)):
                if ok_trans[num] is not None:
                    cache_list=[]
                    for val in ok_trans[num]:
                        cache_list.append(team_cn_list[cn_time])
                        cn_time+=1
                    ok_trans[num]=cache_list
            for i in range(len(need_trans_id)):
                if ok_trans[i] is not None:
                    ok_time=0
                    for im in diff_list[i]:
                        if im['op']=='replace':
                            im['value']=ok_trans[i][ok_time]
                            ok_time+=1
                    final_list.append(jsonpatch.apply_patch(json_for,diff_list[i]))
                else:
                    final_list.append(final_list.append(list_llc[id_llc_true.index(need_trans_id[i])]))
            ret={}
            ret['dataList']=final_list
            with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                json.dump(ret,f,ensure_ascii=False,indent=4)
    else:
        print('???')
        input('按回车继续')
def make_diff_excel(trans_ini,path_llc,path_to,path_for):
        path_llc_list,path_to_list,path_for_list=[],[],[]
        for root, dirs, files in os.walk(path_llc):
            # 遍历当前文件夹下的所有文件
            for file_name in files:
                #file_name=file_name[3:]
                file_path = os.path.join(root, file_name)
                relative_path = os.path.relpath(file_path, start=path_llc)
                path_llc_list.append(relative_path)
        for root, dirs, files in os.walk(path_to):
            # 遍历当前文件夹下的所有文件
            for file_name in files:
                #file_name=file_name[3:]
                file_path = os.path.join(root, file_name)
                relative_path = os.path.relpath(file_path, start=path_to)
                path_to_list.append(relative_path)
        for root, dirs, files in os.walk(path_for):
            # 遍历当前文件夹下的所有文件
            for file_name in files:
                file_name=file_name[3:]
                file_path = os.path.join(root, file_name)
                relative_path = os.path.relpath(file_path, start=path_for)
                path_for_list.append(relative_path)
        for i in path_to:
            if not (i in path_list_jp and i in path_list_en):
                while True:
                    ask=input('文件'+i+'仅在kr文件夹中出现，无法进行常规翻译，是否启用特殊翻译？(y/n)')
                    if ask=='y':
                        one_trans(path_final,llc_path,i,path_output,trans_ini,path_log)
                        break
                    elif ask=='n':
                        name_path=(os.path.split(i))[0]
                        if not name_path=='':
                            name_path=name_path+'\\'
                        name_file=(os.path.split(i))[1]
                        cache=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\en\\'+name_path+'EN_'+name_file
                        mymovefile(cache,path_output,False,name_path)
                        log(path_log,'文件'+i+'仅在kr中出现',trans_ini)
                        break
def cp_trans(path_limbus,path_llc,name,path_output,trans_ini):
    '''
    '''
    name_path=(os.path.split(name))[0]
    if not name_path=='':
        name_path=name_path+'\\'
    name_file=(os.path.split(name))[1]
    path_en=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\en\\'+name_path+'EN_'+name_file
    path_kor=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr\\'+name_path+'KR_'+name_file
    path_jp=path_limbus+'LimbusCompany_Data\Assets\Resources_moved\Localize\jp\\'+name_path+'JP_'+name_file
    path_llc=path_llc+'\\'+name
    if name_file=='E000X.json' :
        print('零协未翻译的愚人节内容')
        if not trans_ini.exc:
            mymovefile(path_en,path_output,True,name_path)
        return 0
    with open(path_en, 'rb') as file:
        json_en=json.load(file) 
    with open(path_kor, 'rb') as file:
        json_kor=json.load(file)
    with open(path_jp, 'rb') as file:
        json_jp=json.load(file)
    llc_exist=True
    if os.path.exists(path_llc):
        with open(path_llc, 'rb') as file:
            json_llc=json.load(file)
    else:llc_exist=False
    #获取列表内容
    if json_en=={}:
        if not trans_ini.exc:
            mymovefile(path_en,path_output,True,name_path)
        return 0
    try:
        list_en=json_en['dataList']
        list_kor=json_kor['dataList']
        list_jp=json_jp['dataList']
    except:
        input('解析'+name+'时出错，不可分析的内容，自动使用原文')
        mymovefile(path_en,path_output,True,name_path)
    if list_en==[] or list_en==[{}]:
        mymovefile(path_en,path_output,True,name_path)
        return 0
    id_en=make_id_list(list_en)
    sheet_name=name.replace('\\','_')
    if trans_ini.exc:
        create_sheet_in_excel(path_output,sheet_name)
    #id_jp=make_id_list(list_jp)
    #id_kor=make_id_list(list_kor)
    if llc_exist:
        list_llc=json_llc['dataList']
        id_llc=make_id_list(list_llc)
        if id_en==id_llc:
            if trans_ini.exc:
                mymovefile(path_llc,path_output,False,name_path)
        else:
            just_none=True
            for i in id_en:
                if not i==None:
                    just_none=False
            if just_none and len(id_en)==len(id_llc):
                if trans_ini.exc:
                    mymovefile(path_llc,path_output,False,name_path)
            else:
                id_llc_true=make_id_list_true(list_llc)
                id_target_true=make_id_list_true(get_json_target(trans_ini.lang_in,json_en,json_kor,json_jp)['dataList'])
                id_for_true=make_id_list_true(get_json_from(trans_ini.lang_in,json_en,json_kor)['dataList'])
                if just_none:
                    print('检测到例外数据，默认未翻译')
                    print('相对路径为'+name)
                    print('绝对路径(零协路径)'+path_llc)
                    print('请检查是否需要翻译')
                    trans_do=''
                    while not (trans_do=='tran' or trans_do=='for' or trans_do=='llc'):
                        print('需要翻译输入tran，使用原文输入for，使用零协输入llc')
                        trans_do=input()
                    if trans_do=='tran':
                        with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                            json.dump(all_trans(get_json_from(trans_ini.lang_in,json_en,json_kor),get_json_target(trans_ini.lang_in,json_en,json_kor,json_jp),trans_ini), f,ensure_ascii=False,indent=4)
                    elif trans_do=='for':
                        if trans_ini.exc:
                            mymovefile(path_en,path_output,True,name_path)
                    else:
                        if trans_ini.exc:
                            mymovefile(path_llc,path_output,False,name_path)
                else:
                    final_list=[]
                    for i in id_target_true:
                        try:
                            id_in_llc=id_llc_true.index(i)+1
                        except:
                            id_in_llc=False
                        if id_in_llc:
                            final_list.append(list_llc[id_in_llc-1])
                        else:
                            json_for=get_json_from(trans_ini.lang_in,json_en,json_kor)['dataList'][id_for_true.index(i)]
                            json_target=get_json_target(trans_ini.lang_in,json_en,json_kor,json_jp)['dataList'][id_target_true.index(i)]
                            diff=jsonpatch.JsonPatch.from_diff(json_for,json_target)
                            for im in diff:
                                if im['op']=='replace':
                                    after=translate_final(trans_ini.lang_in,'zh',im['value'],0,appids,appkeys)
                                    print(after)
                                    im['value']=after
                            final_list.append(jsonpatch.apply_patch(json_for,diff))
                    ret={}
                    ret['dataList']=final_list
                    with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
                        json.dump(ret,f,ensure_ascii=False,indent=4)
    else:
        print(path_en+'    llc文件不存在')
        with open(path_output+'\\'+name,'w',encoding='utf-8') as f:
            json.dump(all_trans(get_json_from(from_lang,json_en,json_kor),get_json_target(from_lang,json_en,json_kor,json_jp),from_lang,0,formal_imput,exc,using_self,cache_trans,team_trans,trans_modu), f,ensure_ascii=False,indent=4)
#from new  import *
    #判断零协是否翻译
#def own_trans

#    if id_en==id_llc:
#        shutil.copy(path_llc,path_output+'\\'+name)
#    else:
#        print(path_en)
#cp_trans('C:\Program Files (x86)\Steam\steamapps\common\Limbus Company\\','C:\Program Files (x86)\Steam\steamapps\common\Limbus Company\LimbusCompany_Data\lang\LLC_CN','StoryData\P11012.json','output\\测试输出','kor',0,'20250405002325015','NMyimh3GSnQm0DUaum1C')
if __name__=='__main_':
    path_lcb='C:/Program Files (x86)/Steam/steamapps/common/Limbus Company/'
    jp,en,kr=get_all_path(path_lcb)
    trans_ini=transform_json()
    key_list=[attr for attr in dir(trans_ini) if not callable(getattr(trans_ini, attr)) and not attr.startswith("__")]
    for i in kr:
        if not i in jp:
            print(i+'  没有对应的日文')
            continue
        ty=find_type(path_lcb,path_lcb+'\\LimbusCompany_Data\\lang\\LLC_zh-CN',i,'a',trans_ini)
        if not (ty=='correct_llc' or ty=='empety'):
            print(i+'  '+ty)
