import winreg
import json
import os
import tkinter as tk
from tkinter import filedialog
#from hashlib import md5
import zipfile
from shutil import rmtree
from shutil import copytree
import requests
import time
import shutil
from openpyxl import load_workbook
from openpyxl import Workbook
from cp_trans_just_list import *
from translate import *
from install import *
from functions import *
import importlib.util
if __name__ == '__main__':
    root = tk.Tk()
    root.withdraw()
class transform_json:
    using_self=False
    cache_trans=False
    lang_in=''
    keys={}
    exc=False
    team_trans=False
    trans_modue=''
    log=False
    json_error=''
def import_module_by_path(module_path):
    """
    根据给定的完整路径动态导入模块
    """
    spec = importlib.util.spec_from_file_location("module_name", module_path)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module
def choose_path():
    f_path = filedialog.askopenfilename()
    return f_path
def find_lcb():
    try:
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER, r'SOFTWARE\Valve\Steam')
        value, value_type = winreg.QueryValueEx(key, 'SteamPath')
        #print(f'Value: {value}, Type: {value_type}')
        winreg.CloseKey(key)
        applist=[]
        with open(value+'\\steamapps\\libraryfolders.vdf', 'r') as f: 
            a=f.read()
            for i in a.split('\n'):
                if 'path' in i:
                    applist.append(i.split('\"')[3])
        for i in applist:
            if os.path.exists(i+'\\steamapps\\common\\Limbus Company\\LimbusCompany.exe'):
                path_lcb=i+'\\steamapps\\common\\Limbus Company\\'
        return path_lcb
    except:None
def has_change():
    try:
       os.remove(os.path.expanduser("~")+'\\limbus.txt')
    except:
        None
    try:
        os.remove('path.txt')
    except:
        None
    exe_name=""
    while not exe_name=="LimbusCompany.exe":
        path_little=choose_path()
        name_list=path_little.split("/")
        exe_name=name_list[-1]
        if not exe_name=="LimbusCompany.exe":
            print("内容错误")
    path_little=path_little[:-17]
    with open(os.path.expanduser("~")+'\\limbus.txt', 'w', encoding='utf-8') as file:
        file.write(path_little)
    with open("path.txt", 'w', encoding='utf-8') as file:
        file.write(path_little)
    return path_little
def write_path(path):
    try:
       os.remove(os.path.expanduser("~")+'\\limbus.txt')
    except:
        None
    try:
        os.remove('path.txt')
    except:
        None
    with open(os.path.expanduser("~")+'\\limbus.txt', 'w', encoding='utf-8') as file:
        file.write(path)
    with open("path.txt", 'w', encoding='utf-8') as file:
        file.write(path)
def find_path():

    if not (os.path.isfile(os.path.expanduser("~")+'\\limbus.txt') and os.path.isfile("path.txt")):
        path_final=find_lcb()
        if path_final is None or (input('这是你的游戏地址吗\n'+path_final+'\nn以否认')=='n'):
            print('选择您的游戏目录')
            input('键入任意键以继续')
            print("请指定游戏路径(选择游戏exe文件)")
            path_final=has_change()
        else:
            write_path(path_final)
    else:
        with open("path.txt", "r") as f:
            data_path = f.readline()
        with open(os.path.expanduser("~")+'\\limbus.txt', "r") as f:
            data_limbus = f.readline()
        if not data_path==data_limbus:
            path_final=find_lcb()
            if path_final is None or (input('这是你的游戏地址吗\n'+path_final+'\nn以否认')=='n'):
                print('选择您的游戏目录')
                input('键入任意键以继续')
                print("请指定游戏路径(选择游戏exe文件)")
                path_final=has_change()
            else:
                write_path(path_final)
        else:
            path_final=data_path
    return path_final
def half_path(trans_ini):
        ask=input('从默认路径寻找吗n以否认')
        if trans_ini.exc:
            if not ask=='n':
                list_now=os.listdir('.')
                can_use=[]
                for i in list_now:
                    if i.endswith('.xlsx'):
                        try:
                            wb = load_workbook(i)
                            if 'init' in wb.sheetnames:can_use.append(i)
                        except:continue
                if len(can_use)==0:
                    print('没有可用的翻译文件')
                    input()
                    raise
                else:print('可使用的翻译文件有：')
                for i in range(len(can_use)):
                    print(str(i+1)+':'+can_use[i])
                while True:
                    choice=input('请输入您要使用的翻译文件序号：')
                    if choice.isdigit() and int(choice)<len(can_use):
                        break
                    else:
                        print('输入错误，请重新输入')
                choice=int(choice)-1
                return can_use[choice]
            else:
                while True:
                    path_cache=choose_path()
                    if path_cache.endswith('.xlsx'):
                        wb = load_workbook(path_cache)
                        if 'init' in wb.sheetnames:
                            return path_cache
                        else:
                            print('请选择正确的文件')
                    else:
                        print('请选择正确的文件')
        else:
            if not ask=='n':
                if not os.path.exists('output'):
                    print('没有output文件夹')
                    input()
                    raise
                list_now=os.listdir('output')
                can_use=[]
                for i in list_now:
                    if os.path.isdir('output\\'+i+'\\Font'):
                        can_use.append(i)
                if len(can_use)==0:
                    print('没有可用的翻译文件')
                    input()
                    raise
                else:print('可使用的翻译文件有：')
                for i in range(len(can_use)):
                    print(str(i+1)+':'+can_use[i])
                while True:
                    choice=input('请输入您要使用的翻译文件序号：')
                    if choice.isdigit() and int(choice)<len(can_use):
                        break
                    else:
                        print('输入错误，请重新输入')
                choice=int(choice)-1
                ret='output\\'+can_use[choice]
                return ret
            else:
                while True:
                    path_cache=filedialog.askdirectory(title="请选择一个文件夹")
                    if os.path.isdir(path_cache+'\\Font'):
                        return path_cache
                    else:
                        print('请选择正确的文件')
None
def find_modul():

    list_now=os.listdir('.')
    list_py=[]
    for i in list_now:
        if i.endswith('.py'):
            list_py.append(i)
    if list_py==[]:
        print('no py file')
        input
        raise
    print('可用py文件列表:')
    for i in range(len(list_py)):
        print(str(i+1)+'.'+list_py[i])
    while True:
        try:
            num=int(input('请输入序号：'))
            if num<len(list_py)+1 and num>0:
                break
            else:
                print('输入错误')
        except:print('输入错误')
    num=num-1
    return list_py[num]
#def check_path(path_jp,path_kr,path_en):
    
None
def get_all_path(path_final):
    path_output_jp=[]
    path_output_en=[]
    path_output_kr=[]
    for root, dirs, files in os.walk(path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\jp'):
        # 遍历当前文件夹下的所有文件
        for file_name in files:
            file_name=file_name[3:]
            file_path = os.path.join(root, file_name)
            relative_path = os.path.relpath(file_path, start=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\jp')
            path_output_jp.append(relative_path)
    for root, dirs, files in os.walk(path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr'):
        # 遍历当前文件夹下的所有文件
        for file_name in files:
            file_name=file_name[3:]
            file_path = os.path.join(root, file_name)
            relative_path = os.path.relpath(file_path, start=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr')
            path_output_kr.append(relative_path)
    for root, dirs, files in os.walk(path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\en'):
        # 遍历当前文件夹下的所有文件
        for file_name in files:
            file_name=file_name[3:]
            file_path = os.path.join(root, file_name)
            relative_path = os.path.relpath(file_path, start=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\en')
            path_output_en.append(relative_path)
    return path_output_jp,path_output_en,path_output_kr
def make_translate(path_final):
    print('请选择零协汉化文件夹(到llc文件夹)')
    llc_path = filedialog.askdirectory(title="请选择一个文件夹") 
    cache=''
    if not (os.path.exists(llc_path+'\\BattleAnnouncerDlg')):
        while not cache=='ok':
            print('错误文件夹,请重新选择')
            llc_path = filedialog.askdirectory(title="请选择一个文件夹") 
            if os.path.exists(llc_path+'\\BattleAnnouncerDlg'):
                cache='ok'
    print('选择您的翻译原语言(仅en支持使用词库(未实现))')
    cache=''
    while not (cache=='kor' or cache=='en' or cache=='jp'):
        print('韩语 kor 英语 en 日文 jp')
        cache=input()
    lang_in=cache
    trans_ini=transform_json(lang_in=lang_in)
    trans_ini.keys={
        #'appids':input('请输入百度翻译appid:'),
        #'appkeys':input('请输入百度翻译appkey:'),
        'appids':'20250405002325015',
        'appkeys':'5zNgTyqclEiOghg70Cx3'
    }
    #print('正在演示，自动填入id')
    #appid = '20250405002325015'
    #appkey = 'NMyimh3GSnQm0DUaum1C'
    #print('正在演示，自动填入key')
    #a=input()
    print('已开始翻译，文件默认放在当前目录output文件夹下')
    path_output='output\\trans'+str(int(time.time()))
    os.makedirs(path_output)
    shutil.copytree(llc_path+'\\Font',path_output+'\\Font')
    path_list=[]
    for root, dirs, files in os.walk(path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr'):
        # 遍历当前文件夹下的所有文件
        for file_name in files:
            file_name=file_name[3:]
            file_path = os.path.join(root, file_name)
            relative_path = os.path.relpath(file_path, start=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\kr')
            path_list.append(relative_path)
    for i in path_list:
        #try:
            print('开始翻译'+i)
            cp_trans(path_final,llc_path,i,path_output,trans_ini)
        #except:
        #    print('翻译'+i+'时出错')
def make_translate_more(path_final):
    print('请选择零协汉化文件夹(到llc文件夹)')
    llc_path = filedialog.askdirectory(title="请选择一个文件夹") 
    cache=''
    if not (os.path.exists(llc_path+'\\BattleAnnouncerDlg')):
        while not cache=='ok':
            print('错误文件夹,请重新选择')
            llc_path = filedialog.askdirectory(title="请选择一个文件夹") 
            if os.path.exists(llc_path+'\\BattleAnnouncerDlg'):
                cache='ok'
    trans_ini=transform_json()
    while True:
        ask=input('是否启用自定义汉化脚本(y/n)(?以帮助):')
        if ask=='y' or ask=='n':
            break
        elif ask=='？' or ask=='?':
            print('启用自定义汉化脚本功能，将会在从当前目录中导入self_trans.py脚本，其中包含函数trans_self(input)，返回汉化完成后的内容')
            print('详情参考example.py脚本')
            print('脚本中所用api或许需要进行预处理以防止翻译[]和<>中的内容')
            print('脚本将在运行时被导入，请确保脚本中没有语法错误')
            print('脚本如需初始化，请将初始化函数放在inits()中')
            print('tips:你可以通过选择just_formed.py搭配excel设置以输出翻译原文')
        else:
            print('输入错误')
    if ask=='y':
        trans_ini.using_self=True
    else:
        trans_ini.using_self=False
    while True:
        print('是否启用实验室功能(?以帮助):')
        print('1.缓存翻译结果')
        print('2.整文件翻译')
        print('n以拒绝')
        ask=input('输入数字或?以帮助:')
        if ask=='1' or ask=='2' or ask=='n':
            break
        elif ask=='？' or ask=='?':
            print('1.缓存翻译结果:将翻译结果缓存，后续相同内容时直接调用以节约api调用次数')
            print('2.整文件翻译:将整文件中翻译内容整合后丢入翻译脚本，以减少翻译次数，以列表形式输入，要求输出同样长度的列表，否则会报错停机(必须启用自定义脚本功能且与缓存功能冲突)')
        else:
            print('输入错误')
    if ask=='1':
        trans_ini.cache_trans=True
    elif ask=='2':
        trans_ini.team_trans=True
    while True:
        ask=input('是否将修改输出至excel文件?(y/n)?以帮助')
        if ask=='y' or ask=='n':
            break
        elif ask=='?' or ask=='？':
            print('将修改输出至excel文件，允许用户进行修改后apply')
    if ask=='y':
        trans_ini.exc=True
    else:
        trans_ini.exc=False
    while True:
        ask=input('是否启用log(y/n)')
        if ask=='y':
            trans_ini.log=True
            break
        elif ask=='n':
            trans_ini.log=False
            break
        else:print('输入错误')
    ask=input('是否使用翻译至一半的项目?仅y以确认')
    if ask=='y':
        half_trans=True
    else:
        half_trans=False
    half_trans_path=[]
    if half_trans and (not trans_ini.exc):
        #获取翻译至一半的项目(dirs)
        half_trans_path=half_path(trans_ini)
        half_trans_list=[]
        for root, dirs, files in os.walk(half_trans_path):
        # 遍历当前文件夹下的所有文件
            for file_name in files:
                file_path = os.path.join(root, file_name)
                relative_path = os.path.relpath(file_path, start=half_trans_path)
                half_trans_list.append(relative_path)
    elif half_trans:
        # 读取翻译一半的文件
        excel_read=read_sheet_content(half_trans_path, 'have_trans')
        half_trans_list=[]
        for i in excel_read:
            half_trans_list.append(i[0])
        shutil.copy(half_trans_path,path_output)
    if True:
        formal_input=None
        if trans_ini.using_self:
            path_modul=find_modul()
            try:
                trans_ini.trans_modue=import_module_by_path(path_modul)
            except:
                print('加载脚本失败')
                input()
                raise
            if not 'trans_selfdir' in (trans_ini.trans_modue):
                print('脚本缺少trans_selfdir()函数')
                input()
                raise
            if not 'inits' in (trans_ini.trans_modue):
                print('警告:脚本缺少inits()函数')
                input('按回车继续')
            else:
                try:
                    trans_ini.keys=trans_ini.trans_modue.inits()
                except:input('警告:inits()函数执行错误')
        else:
            trans_ini.keys={
                #'appids':input('请输入百度翻译appid:'),
                #'appkeys':input('请输入百度翻译appkey:'),
                'appids':'20250405002325015',
                'appkeys':'5zNgTyqclEiOghg70Cx3'
            }
        print('选择您的翻译原语言(仅en支持使用词库(未实现))')
        cache=''
        while not (cache=='kor' or cache=='en' or cache=='jp'):
            print('韩语 kor 英语 en 日文 jp')
            cache=input()
        trans_ini.lang_in=cache
        if not trans_ini.exc:
            print('已开始翻译，文件默认放在当前目录output文件夹下')
        else:print('已开始翻译，文件默认放在当前目录output表格下')
        time_now=str(int(time.time()))
        if (not trans_ini.exc) and (not half_trans):
            path_output=('output\\trans'+time_now)
            os.makedirs(path_output)
            shutil.copytree(llc_path+'\\Font',path_output+'\\Font')
        else:
            path_output=('output'+time_now+'.xlsx')
            create_sheet_in_excel(path_output, 'init')
            create_sheet_in_excel(path_output, 'have_trans')
        path_log='log\\trans'+time_now+'.log'
        try:
            os.makedirs('log')
        except:pass
        with open(path_log,'a') as f:
            f.write('开始翻译\n时间戳:'+time_now+'\nllc路径:'+llc_path+'\nlcb路径'+path_final+'\n输出文件:'+path_output+'\n')
        path_list_jp,path_list_en,path_list_kr=get_all_path(path_final)
        for i in path_list_kr:
            if half_trans:
                if i in half_trans_list:
                    log(path_log,i+'已经完成翻译',trans_ini)
                    if not trans_ini.exc:
                        shutil.copy(half_trans_path+'\\'+i,path_output+'\\'+i)
                    continue
            if not (i in path_list_jp and i in path_list_en):
                while True:
                    ask=input('文件'+i+'仅在kr文件夹中出现，无法进行常规翻译，是否启用特殊翻译？(y/n)')
                    if ask=='y':
                        one_trans(path_final,llc_path,i,path_output,trans_ini,path_log)
                        break
                    elif ask=='n':
                        name_path=(os.path.split(i))[0]
                        if not name_path=='':
                            name_path=name_path+'\\'
                        name_file=(os.path.split(i))[1]
                        cache=path_final+'LimbusCompany_Data\Assets\Resources_moved\Localize\en\\'+name_path+'EN_'+name_file
                        mymovefile(cache,path_output,False,name_path)
                        log(path_log,'文件'+i+'仅在kr中出现',trans_ini)
                        break
            try:
                print('开始翻译'+i)
                cp_trans_new(path_final,llc_path,i,path_output,trans_ini,path_log)
            except Exception as e:
                print('翻译'+i+'时出错\nerrorcode'+str(e))
                log(path_log,'翻译'+i+'时出错\nerrorcode'+str(e),trans_ini)
                if input()=='y':
                    cp_trans_new(path_final,llc_path,i,path_output,trans_ini,path_log)
if __name__ == '__main__':
    path_final=find_path()
    while True:
        print('请选择操作:')
        print('1. 使用自动汉化')
        print('2. 启用高级操作')
        print('3. 启用其他工具')
        print('?以帮助')
        print('new.LCTA项目简介')
        print('up.本次更新内容')
        ask=input('输入数字或?以帮助:')
        if ask=='1' or ask=='2' or ask=='3':
            break
        elif ask=='？' or ask=='?':
            print('1.在不启用高级设置的情况下进行汉化，汉化速度高，且bug少(即LCTA1.0原版)')
            print('2.在启用高级设置的情况下进行汉化，高级功能包括{\n启用自定义翻译脚本\n缓存翻译内容\n从完成一半的汉化文件中继续\n整文件统一翻译\n将修改输出至excel文件}\n速度较慢,bug较多')
            print('3.使用LCTA的其他工具，包括将本地文件进行安装，根据上文所说的excel文件进行汉化')
        elif ask=='new':
            print('LCTA即LimbusCompany Translate Automation，是一项开源边狱公司自动翻译软件')
            print('当前版本2.0.0')
            print('github开源链接HZBHZB1234/LCTA-Limbus-company-transfer-auto')
            print('使用本软件产出的项目需要在引用时标明')
            print('作者不承担任何因本软件造成的损失的责任')
            print('作者贴吧:HZBHZB31415926')
            print('作者github:HZBHZB1234')
            print('作者bilibili:ygdtpnn')
            print('如软件出现任何bug可以在github提issue或是在贴吧或b站回复')
        elif ask=='up':
            print('2.0.0版本更新内容:\n添加了高级功能库\n添加了工具箱功能\n修复了在翻译仅有韩语版本的文本时出现的错误\n修复了少量bug\n增加了更多bug')
        else:
            print('输入错误')
    if ask=='1':
        make_translate(path_final)
    if ask=='2':
        make_translate_more(path_final)
    if ask=='3':
        while True:
            ask=input('1.安装已有汉化\n2.使用excel表格附加\n3.从ourplay加速器下载现成汉化\n4.清除本地配置文件\n5.从零协下载汉化\n?以帮助')
            if ask=='1':
                install()
            elif ask=='2':
                None
            elif ask=='3':
                url,md5s,size=download_ourplay()
                print(f'下载地址：{url}\nMD5：{md5s}\n大小：{size}')
                input('回车以确认下载')
                if not download_and_verify_file():
                    continue
                ask=input('是否格式化文件以支持调用？(n以跳过)')
                if not ask=='n':
                    try:
                        get_true_zip('transfile.zip')
                    except Exception as e:
                        print(f'格式化文件时出现错误，错误信息:{e}')
                        input('回车以清除中间文件')
                        try:shutil.rmtree('temp')
                        except:None
                        try:os.remove('transfile.zip')
                        except:None
            elif ask=='4':
                cache_dir=os.path.expanduser('~')+'\\AppData\\LocalLow\\ProjectMoon\\LimbusCompany'
                ask=input(f'这是你的缓存文件夹吗n以否认\n{cache_dir}')
                if ask=='n':
                    cache_dir=input('请输入缓存文件夹路径')
                if not os.path.exists(cache_dir+'\\Assets'):
                    input('请检查路径是否正确')
                    break
                files=os.listdir(cache_dir)
                print('开始清理')
                for file in files:
                    if not file=='Assets':
                        print('删除'+file)
                        if os.path.isdir(cache_dir+'\\'+file):
                            shutil.rmtree(cache_dir+'\\'+file)
                        else:
                            os.remove(cache_dir+'\\'+file)
            elif ask=='5':
                download_llc()
            elif ask=='?' or ask=='？':
                print('1.安装已有汉化(放在当前文件夹)')
                print('2.使用excel表格附加(基于高级翻译功能生成的excel表格)')
                print('3.从ourplay加速器下载现成汉化(建议转化为LCTA通用格式)(建议搭配其余功能使用)')
                print('4.清除缓存的本地配置文件，可以用于恢复默认设置，牢六更新时留')
                print('5.下载零协汉化，可转化为LCTA通用格式，建议搭配其他功能使用')
            else:
                print('输入错误')
if __name__=='__main__':
    print(find_modul())